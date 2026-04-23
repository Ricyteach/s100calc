#!/usr/bin/env python3
"""
scripts/export_esr.py

Export the ESR-2093 section properties sheet from the AISI S100 Hat Section
calculator workbook into a CSV file consumable by the s100calc ESR registry.

The output CSV has a single header row (sheet row 5) followed by one row per
section (sheet rows 6 through the last row with a populated `Key` cell).
Empty cells are emitted as empty fields. Footnote rows below the data block
are excluded.

Runtime schema checks fail fast if the workbook layout has drifted:
- Sheet named ESR-2093 exists
- Row 5 column A is the literal 'Key'
- Row 5 column B is the literal 'TradeName'
- Every data row has a non-empty Key value

Usage:
    python scripts/export_esr.py --workbook path/to/workbook.xlsx
    python scripts/export_esr.py \\
        --workbook AISI_S100_HatSection_Calc_v10_2_de365.xlsx \\
        --output tests/esr/esr_2093/esr_2093.csv
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

SHEET_NAME = "ESR-2093"
HEADER_ROW = 5
FIRST_DATA_ROW = 6


def export(workbook_path: Path, output_path: Path) -> tuple[int, int]:
    """Read the ESR-2093 sheet and write it to ``output_path``. Returns (rows, cols)."""
    if not workbook_path.is_file():
        sys.exit(f"error: workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"error: sheet {SHEET_NAME!r} not found in {workbook_path.name}")

    ws = wb[SHEET_NAME]
    ncols = ws.max_column

    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ncols + 1)]
    if headers[0] != "Key" or headers[1] != "TradeName":
        sys.exit(
            f"error: row {HEADER_ROW} does not start with ('Key', 'TradeName'): "
            f"found ({headers[0]!r}, {headers[1]!r}). Sheet schema may have drifted."
        )

    data_rows: list[list] = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if key is None:
            break  # blank row: end of data, start of footnotes block
        data_rows.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])

    if not data_rows:
        sys.exit("error: no data rows found below header")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data_rows)

    return len(data_rows), ncols


def main() -> int:
    default_output = (
        Path(__file__).resolve().parent.parent
        / "tests" / "esr" / "esr_2093" / "esr_2093.csv"
    )

    p = argparse.ArgumentParser(
        description="Export ESR-2093 section properties sheet to registry-consumable CSV.",
    )
    p.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="path to the de365 .xlsx workbook with the ESR-2093 sheet",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help=f"output CSV path (default: {default_output.relative_to(default_output.parents[3])})",
    )
    args = p.parse_args()

    nrows, ncols = export(args.workbook, args.output)
    size = args.output.stat().st_size
    print(f"exported {nrows} sections x {ncols} columns")
    print(f"wrote {args.output} ({size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())

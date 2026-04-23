#!/usr/bin/env python3
"""
scripts/export_oracle.py

Export the TEST_VECTORS sheet from the AISI S100 Hat Section calculator workbook
into a pytest-consumable CSV oracle.

The output CSV has a single header row (sheet row 5) followed by one row per
captured case (sheet rows 6 through the last row with a populated `id` cell).
Empty cells are emitted as empty fields; pytest code asserts only on non-blank
columns per row (see the NOTES section of the TEST_VECTORS sheet).

Runtime schema checks fail fast if the workbook layout has drifted:
- Sheet named TEST_VECTORS exists
- Row 5 column A is the literal 'id'
- Row 5 column B is the literal 'slug'
- Data rows have sequential ids starting at 1

Usage:
    python scripts/export_oracle.py --workbook path/to/workbook.xlsx
    python scripts/export_oracle.py \\
        --workbook AISI_S100_HatSection_Calc_v10_2_with_tests.xlsx \\
        --output tests/esr/esr_2093/oracle.csv
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

SHEET_NAME = "TEST_VECTORS"
HEADER_ROW = 5
FIRST_DATA_ROW = 6


def export(workbook_path: Path, output_path: Path) -> tuple[int, int]:
    """Read TEST_VECTORS and write it to ``output_path``. Returns (rows, cols)."""
    if not workbook_path.is_file():
        sys.exit(f"error: workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"error: sheet {SHEET_NAME!r} not found in {workbook_path.name}")

    ws = wb[SHEET_NAME]
    ncols = ws.max_column

    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ncols + 1)]
    if headers[0] != "id" or headers[1] != "slug":
        sys.exit(
            f"error: row {HEADER_ROW} does not start with ('id', 'slug'): "
            f"found ({headers[0]!r}, {headers[1]!r}). Sheet schema may have drifted."
        )

    data_rows: list[list] = []
    expected_id = 1
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        case_id = ws.cell(row=r, column=1).value
        if case_id is None:
            break  # blank row: end of data, start of NOTES block
        if case_id != expected_id:
            sys.exit(
                f"error: row {r} id is {case_id!r}, expected {expected_id} "
                f"(case id sequence broken)"
            )
        data_rows.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])
        expected_id += 1

    if not data_rows:
        sys.exit("error: no data rows found below header")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data_rows)

    return len(data_rows), ncols


def main() -> int:
    default_output = (
        Path(__file__).resolve().parent.parent
        / "tests" / "esr" / "esr_2093" / "oracle.csv"
    )

    p = argparse.ArgumentParser(
        description="Export TEST_VECTORS sheet to pytest-consumable CSV oracle.",
    )
    p.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="path to the de365 .xlsx workbook with the TEST_VECTORS sheet",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=default_output,
        help=f"output CSV path (default: {default_output.relative_to(default_output.parents[3])})",
    )
    args = p.parse_args()

    nrows, ncols = export(args.workbook, args.output)
    size = args.output.stat().st_size
    print(f"exported {nrows} cases x {ncols} columns")
    print(f"wrote {args.output} ({size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
